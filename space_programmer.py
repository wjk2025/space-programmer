#!/usr/bin/env python3
"""
Architecture Space Programming Tool
Calculates space requirements based on staff counts, department allocations,
and industry-standard factors for circulation and rentable conversion.
"""

import json
import os
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional
from pathlib import Path


# ============================================================================
# CONFIGURATION - Industry Standard Space Standards
# ============================================================================

DEFAULT_SPACE_STANDARDS = {
    "workstation_open": {"name": "Open Workstation", "sf": 48, "description": "Open plan workstation"},
    "workstation_standard": {"name": "Standard Workstation", "sf": 64, "description": "Standard cubicle workstation"},
    "workstation_large": {"name": "Large Workstation", "sf": 80, "description": "Large/senior workstation"},
    "office_small": {"name": "Small Private Office", "sf": 100, "description": "Small enclosed office"},
    "office_standard": {"name": "Standard Private Office", "sf": 150, "description": "Standard private office"},
    "office_large": {"name": "Large Private Office", "sf": 200, "description": "Executive/large private office"},
    "office_executive": {"name": "Executive Suite", "sf": 300, "description": "C-suite/executive office"},
    "conference_small": {"name": "Small Conference Room", "sf": 150, "description": "4-6 person meeting room"},
    "conference_medium": {"name": "Medium Conference Room", "sf": 300, "description": "8-12 person conference room"},
    "conference_large": {"name": "Large Conference Room", "sf": 500, "description": "16-20 person board room"},
    "huddle_room": {"name": "Huddle Room", "sf": 80, "description": "2-4 person huddle space"},
    "phone_booth": {"name": "Phone Booth", "sf": 35, "description": "Single person phone/focus booth"},
    "break_room": {"name": "Break Room", "sf": 200, "description": "Kitchen/break area"},
    "reception": {"name": "Reception Area", "sf": 250, "description": "Reception/waiting area"},
    "copy_print": {"name": "Copy/Print Center", "sf": 100, "description": "Copy/print/mail area"},
    "storage": {"name": "Storage Room", "sf": 150, "description": "General storage"},
    "server_room": {"name": "Server/IT Room", "sf": 120, "description": "Server/IT closet"},
    "wellness_room": {"name": "Wellness/Mother's Room", "sf": 80, "description": "Wellness/lactation room"},
    "training_room": {"name": "Training Room", "sf": 600, "description": "Large training/all-hands room"},
    "collaboration_area": {"name": "Collaboration Zone", "sf": 200, "description": "Open collaboration space"},
}

# Industry standard factors
DEFAULT_CIRCULATION_FACTOR = 0.35  # 35% added for circulation
DEFAULT_LOSS_FACTOR = 0.15  # 15% loss factor for USF to RSF conversion

# Remote work reduction factors
REMOTE_WORK_FACTORS = {
    "full_onsite": {"factor": 1.0, "description": "100% on-site, no reduction"},
    "hybrid_light": {"factor": 0.85, "description": "1 day remote/week (~15% reduction)"},
    "hybrid_moderate": {"factor": 0.70, "description": "2 days remote/week (~30% reduction)"},
    "hybrid_heavy": {"factor": 0.55, "description": "3 days remote/week (~45% reduction)"},
    "remote_first": {"factor": 0.40, "description": "4 days remote/week (~60% reduction)"},
    "fully_remote": {"factor": 0.15, "description": "Fully remote, hoteling only (~85% reduction)"},
}


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class Department:
    """Represents a department with staff allocations"""
    name: str
    open_workstations: int = 0
    standard_workstations: int = 0
    large_workstations: int = 0
    small_offices: int = 0
    standard_offices: int = 0
    large_offices: int = 0
    executive_offices: int = 0
    
    @property
    def total_staff(self) -> int:
        return (self.open_workstations + self.standard_workstations + 
                self.large_workstations + self.small_offices + 
                self.standard_offices + self.large_offices + 
                self.executive_offices)


@dataclass
class SupportSpaces:
    """Common support/amenity spaces"""
    small_conference: int = 0
    medium_conference: int = 0
    large_conference: int = 0
    huddle_rooms: int = 0
    phone_booths: int = 0
    break_rooms: int = 0
    reception_areas: int = 0
    copy_print_centers: int = 0
    storage_rooms: int = 0
    server_rooms: int = 0
    wellness_rooms: int = 0
    training_rooms: int = 0
    collaboration_areas: int = 0


@dataclass
class SpaceProgram:
    """Complete space program with all inputs and calculations"""
    # Company information
    company_name: str = ""
    location: str = ""
    project_name: str = ""
    prepared_by: str = ""
    date_created: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))
    
    # Departments
    departments: list = field(default_factory=list)
    
    # Support spaces
    support_spaces: SupportSpaces = field(default_factory=SupportSpaces)
    
    # Factors
    circulation_factor: float = DEFAULT_CIRCULATION_FACTOR
    loss_factor: float = DEFAULT_LOSS_FACTOR
    remote_work_policy: str = "full_onsite"
    
    # Custom space standards (if overriding defaults)
    custom_standards: dict = field(default_factory=dict)
    
    # Notes
    notes: str = ""
    
    def get_space_standard(self, key: str) -> dict:
        """Get space standard, checking custom first then defaults"""
        if key in self.custom_standards:
            return self.custom_standards[key]
        return DEFAULT_SPACE_STANDARDS.get(key, {"sf": 0})


# ============================================================================
# CALCULATION ENGINE
# ============================================================================

class SpaceCalculator:
    """Performs all space calculations"""
    
    def __init__(self, program: SpaceProgram):
        self.program = program
    
    def calculate_department_sf(self, dept: Department) -> dict:
        """Calculate square footage for a department"""
        p = self.program
        breakdown = {
            "open_workstations": dept.open_workstations * p.get_space_standard("workstation_open")["sf"],
            "standard_workstations": dept.standard_workstations * p.get_space_standard("workstation_standard")["sf"],
            "large_workstations": dept.large_workstations * p.get_space_standard("workstation_large")["sf"],
            "small_offices": dept.small_offices * p.get_space_standard("office_small")["sf"],
            "standard_offices": dept.standard_offices * p.get_space_standard("office_standard")["sf"],
            "large_offices": dept.large_offices * p.get_space_standard("office_large")["sf"],
            "executive_offices": dept.executive_offices * p.get_space_standard("office_executive")["sf"],
        }
        breakdown["total"] = sum(breakdown.values())
        return breakdown
    
    def calculate_support_sf(self) -> dict:
        """Calculate square footage for support spaces"""
        p = self.program
        ss = p.support_spaces
        breakdown = {
            "small_conference": ss.small_conference * p.get_space_standard("conference_small")["sf"],
            "medium_conference": ss.medium_conference * p.get_space_standard("conference_medium")["sf"],
            "large_conference": ss.large_conference * p.get_space_standard("conference_large")["sf"],
            "huddle_rooms": ss.huddle_rooms * p.get_space_standard("huddle_room")["sf"],
            "phone_booths": ss.phone_booths * p.get_space_standard("phone_booth")["sf"],
            "break_rooms": ss.break_rooms * p.get_space_standard("break_room")["sf"],
            "reception_areas": ss.reception_areas * p.get_space_standard("reception")["sf"],
            "copy_print_centers": ss.copy_print_centers * p.get_space_standard("copy_print")["sf"],
            "storage_rooms": ss.storage_rooms * p.get_space_standard("storage")["sf"],
            "server_rooms": ss.server_rooms * p.get_space_standard("server_room")["sf"],
            "wellness_rooms": ss.wellness_rooms * p.get_space_standard("wellness_room")["sf"],
            "training_rooms": ss.training_rooms * p.get_space_standard("training_room")["sf"],
            "collaboration_areas": ss.collaboration_areas * p.get_space_standard("collaboration_area")["sf"],
        }
        breakdown["total"] = sum(breakdown.values())
        return breakdown
    
    def calculate_totals(self) -> dict:
        """Calculate all totals with factors applied"""
        # Department totals
        dept_results = []
        total_dept_sf = 0
        total_staff = 0
        
        for dept in self.program.departments:
            dept_sf = self.calculate_department_sf(dept)
            dept_results.append({
                "name": dept.name,
                "staff": dept.total_staff,
                "breakdown": dept_sf,
            })
            total_dept_sf += dept_sf["total"]
            total_staff += dept.total_staff
        
        # Support space totals
        support_sf = self.calculate_support_sf()
        
        # Net assignable SF (before circulation)
        net_assignable_sf = total_dept_sf + support_sf["total"]
        
        # Usable SF (with circulation)
        circulation_sf = net_assignable_sf * self.program.circulation_factor
        usable_sf = net_assignable_sf + circulation_sf
        
        # Remote work adjustment
        remote_factor = REMOTE_WORK_FACTORS[self.program.remote_work_policy]["factor"]
        adjusted_usable_sf = usable_sf * remote_factor
        
        # Rentable SF (with loss factor)
        loss_sf = adjusted_usable_sf * self.program.loss_factor
        rentable_sf = adjusted_usable_sf + loss_sf
        
        # Calculate SF per person metrics
        sf_per_person_net = net_assignable_sf / total_staff if total_staff > 0 else 0
        sf_per_person_usable = usable_sf / total_staff if total_staff > 0 else 0
        sf_per_person_adjusted = adjusted_usable_sf / total_staff if total_staff > 0 else 0
        sf_per_person_rentable = rentable_sf / total_staff if total_staff > 0 else 0
        
        return {
            "company": {
                "name": self.program.company_name,
                "location": self.program.location,
                "project_name": self.program.project_name,
                "prepared_by": self.program.prepared_by,
                "date": self.program.date_created,
            },
            "departments": dept_results,
            "support_spaces": support_sf,
            "totals": {
                "total_staff": total_staff,
                "department_sf": total_dept_sf,
                "support_sf": support_sf["total"],
                "net_assignable_sf": net_assignable_sf,
                "circulation_factor": self.program.circulation_factor,
                "circulation_sf": circulation_sf,
                "usable_sf": usable_sf,
                "remote_work_policy": self.program.remote_work_policy,
                "remote_work_description": REMOTE_WORK_FACTORS[self.program.remote_work_policy]["description"],
                "remote_adjustment_factor": remote_factor,
                "adjusted_usable_sf": adjusted_usable_sf,
                "loss_factor": self.program.loss_factor,
                "loss_sf": loss_sf,
                "rentable_sf": rentable_sf,
            },
            "metrics": {
                "sf_per_person_net": sf_per_person_net,
                "sf_per_person_usable": sf_per_person_usable,
                "sf_per_person_adjusted": sf_per_person_adjusted,
                "sf_per_person_rentable": sf_per_person_rentable,
            },
            "notes": self.program.notes,
        }


# ============================================================================
# REMOTE WORK ANALYZER
# ============================================================================

class RemoteWorkAnalyzer:
    """Analyzes impact of remote work policies on space requirements"""
    
    def __init__(self, program: SpaceProgram):
        self.program = program
        self.calculator = SpaceCalculator(program)
    
    def analyze_scenarios(self) -> dict:
        """Analyze all remote work scenarios and their impact"""
        # Get base calculation (full onsite)
        original_policy = self.program.remote_work_policy
        self.program.remote_work_policy = "full_onsite"
        base_results = self.calculator.calculate_totals()
        base_usable = base_results["totals"]["usable_sf"]
        base_rentable = base_results["totals"]["rentable_sf"]
        
        scenarios = []
        for policy_key, policy_info in REMOTE_WORK_FACTORS.items():
            self.program.remote_work_policy = policy_key
            results = self.calculator.calculate_totals()
            
            adjusted_usable = results["totals"]["adjusted_usable_sf"]
            adjusted_rentable = results["totals"]["rentable_sf"]
            
            scenarios.append({
                "policy": policy_key,
                "description": policy_info["description"],
                "reduction_factor": 1 - policy_info["factor"],
                "adjusted_usable_sf": adjusted_usable,
                "adjusted_rentable_sf": adjusted_rentable,
                "usable_sf_saved": base_usable - adjusted_usable,
                "rentable_sf_saved": base_rentable - adjusted_rentable,
                "percent_reduction": (1 - policy_info["factor"]) * 100,
            })
        
        # Restore original policy
        self.program.remote_work_policy = original_policy
        
        return {
            "base_usable_sf": base_usable,
            "base_rentable_sf": base_rentable,
            "scenarios": scenarios,
            "recommendations": self._generate_recommendations(scenarios, base_results["totals"]["total_staff"]),
        }
    
    def _generate_recommendations(self, scenarios: list, total_staff: int) -> list:
        """Generate recommendations based on analysis"""
        recommendations = []
        
        recommendations.append({
            "category": "Hoteling Strategy",
            "text": f"For hybrid policies, consider implementing desk hoteling. "
                   f"With {total_staff} staff and a hybrid-moderate policy (2 days remote), "
                   f"you could reduce dedicated desks to approximately {int(total_staff * 0.7)} stations."
        })
        
        recommendations.append({
            "category": "Meeting Space",
            "text": "With increased remote work, consider reallocating saved space to "
                   "additional video-conference enabled meeting rooms and collaboration zones "
                   "to support hybrid meetings."
        })
        
        recommendations.append({
            "category": "Amenity Enhancement",
            "text": "Remote work policies often work best when office amenities are enhanced. "
                   "Consider adding: focus/phone booths, wellness rooms, and upgraded break areas "
                   "to make in-office days more productive."
        })
        
        return recommendations


# ============================================================================
# DATA PERSISTENCE
# ============================================================================

class DataManager:
    """Handles saving and loading program data"""
    
    def __init__(self, data_dir: str = "./data"):
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
    
    def save_program(self, program: SpaceProgram, filename: Optional[str] = None) -> str:
        """Save program to JSON file"""
        if filename is None:
            safe_name = "".join(c if c.isalnum() else "_" for c in program.company_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{safe_name}_{timestamp}.json"
        
        filepath = self.data_dir / filename
        
        # Convert to serializable format
        data = {
            "company_name": program.company_name,
            "location": program.location,
            "project_name": program.project_name,
            "prepared_by": program.prepared_by,
            "date_created": program.date_created,
            "departments": [asdict(d) if isinstance(d, Department) else d for d in program.departments],
            "support_spaces": asdict(program.support_spaces) if isinstance(program.support_spaces, SupportSpaces) else program.support_spaces,
            "circulation_factor": program.circulation_factor,
            "loss_factor": program.loss_factor,
            "remote_work_policy": program.remote_work_policy,
            "custom_standards": program.custom_standards,
            "notes": program.notes,
        }
        
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2)
        
        return str(filepath)
    
    def load_program(self, filename: str) -> SpaceProgram:
        """Load program from JSON file"""
        filepath = self.data_dir / filename
        
        with open(filepath, 'r') as f:
            data = json.load(f)
        
        # Convert departments back to objects
        departments = [Department(**d) for d in data.get("departments", [])]
        
        # Convert support spaces back to object
        support_data = data.get("support_spaces", {})
        support_spaces = SupportSpaces(**support_data) if support_data else SupportSpaces()
        
        program = SpaceProgram(
            company_name=data.get("company_name", ""),
            location=data.get("location", ""),
            project_name=data.get("project_name", ""),
            prepared_by=data.get("prepared_by", ""),
            date_created=data.get("date_created", ""),
            departments=departments,
            support_spaces=support_spaces,
            circulation_factor=data.get("circulation_factor", DEFAULT_CIRCULATION_FACTOR),
            loss_factor=data.get("loss_factor", DEFAULT_LOSS_FACTOR),
            remote_work_policy=data.get("remote_work_policy", "full_onsite"),
            custom_standards=data.get("custom_standards", {}),
            notes=data.get("notes", ""),
        )
        
        return program
    
    def list_programs(self) -> list:
        """List all saved programs"""
        files = list(self.data_dir.glob("*.json"))
        programs = []
        for f in files:
            try:
                with open(f, 'r') as file:
                    data = json.load(file)
                programs.append({
                    "filename": f.name,
                    "company_name": data.get("company_name", "Unknown"),
                    "location": data.get("location", ""),
                    "date_created": data.get("date_created", ""),
                })
            except:
                pass
        return programs


# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_to_excel(results: dict, remote_analysis: dict, program: SpaceProgram, output_path: str):
    """Export results to Excel workbook"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="2E5090")
    subheader_fill = PatternFill("solid", fgColor="4472C4")
    section_fill = PatternFill("solid", fgColor="B4C6E7")
    total_fill = PatternFill("solid", fgColor="FFC000")
    input_font = Font(color="0000FF")  # Blue for inputs
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ========== SUMMARY SHEET ==========
    ws = wb.active
    ws.title = "Summary"
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = "SPACE PROGRAMMING SUMMARY"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Company Info
    row = 3
    info_items = [
        ("Company:", results["company"]["name"]),
        ("Location:", results["company"]["location"]),
        ("Project:", results["company"]["project_name"]),
        ("Prepared By:", results["company"]["prepared_by"]),
        ("Date:", results["company"]["date"]),
    ]
    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # Space Summary
    row += 2
    ws[f'A{row}'] = "SPACE SUMMARY"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    summary_items = [
        ("Total Headcount", results["totals"]["total_staff"], ""),
        ("Department Space", results["totals"]["department_sf"], "SF"),
        ("Support Space", results["totals"]["support_sf"], "SF"),
        ("Net Assignable SF", results["totals"]["net_assignable_sf"], "SF"),
        (f"Circulation ({results['totals']['circulation_factor']*100:.0f}%)", results["totals"]["circulation_sf"], "SF"),
        ("Usable Square Feet", results["totals"]["usable_sf"], "SF"),
        (f"Remote Work Adjustment ({results['totals']['remote_work_description']})", "", ""),
        ("Adjusted Usable SF", results["totals"]["adjusted_usable_sf"], "SF"),
        (f"Loss Factor ({results['totals']['loss_factor']*100:.0f}%)", results["totals"]["loss_sf"], "SF"),
        ("RENTABLE SQUARE FEET", results["totals"]["rentable_sf"], "SF"),
    ]
    
    for label, value, unit in summary_items:
        ws[f'A{row}'] = label
        if value:
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = unit
        if "RENTABLE" in label:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = total_fill
            ws[f'B{row}'].fill = total_fill
            ws[f'C{row}'].fill = total_fill
        row += 1
    
    # Metrics
    row += 2
    ws[f'A{row}'] = "KEY METRICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    metrics = [
        ("SF per Person (Net)", results["metrics"]["sf_per_person_net"]),
        ("SF per Person (Usable)", results["metrics"]["sf_per_person_usable"]),
        ("SF per Person (Adjusted)", results["metrics"]["sf_per_person_adjusted"]),
        ("SF per Person (Rentable)", results["metrics"]["sf_per_person_rentable"]),
    ]
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = '#,##0.0'
        ws[f'C{row}'] = "SF"
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    
    # ========== DEPARTMENT DETAIL SHEET ==========
    ws2 = wb.create_sheet("Department Detail")
    
    headers = ["Department", "Staff", "Open WS", "Std WS", "Large WS", 
               "Sm Office", "Std Office", "Lg Office", "Exec Office", "Total SF"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row = 2
    for dept in results["departments"]:
        d = next((x for x in program.departments if x.name == dept["name"]), None)
        if d:
            values = [dept["name"], dept["staff"], d.open_workstations, d.standard_workstations,
                     d.large_workstations, d.small_offices, d.standard_offices, 
                     d.large_offices, d.executive_offices, dept["breakdown"]["total"]]
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                if col > 1:
                    cell.number_format = '#,##0'
            row += 1
    
    # Totals row
    total_row = row
    ws2[f'A{row}'] = "TOTAL"
    ws2[f'A{row}'].font = Font(bold=True)
    for col in range(2, 11):
        cell = ws2.cell(row=row, column=col)
        col_letter = get_column_letter(col)
        cell.value = f'=SUM({col_letter}2:{col_letter}{row-1})'
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.number_format = '#,##0'
    
    for col in range(1, 11):
        ws2.column_dimensions[get_column_letter(col)].width = 12 if col > 1 else 25
    
    # ========== SUPPORT SPACES SHEET ==========
    ws3 = wb.create_sheet("Support Spaces")
    
    support_headers = ["Space Type", "Quantity", "SF Each", "Total SF"]
    for col, header in enumerate(support_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    support_items = [
        ("Small Conference Room (4-6)", program.support_spaces.small_conference, 150),
        ("Medium Conference Room (8-12)", program.support_spaces.medium_conference, 300),
        ("Large Conference Room (16-20)", program.support_spaces.large_conference, 500),
        ("Huddle Rooms (2-4)", program.support_spaces.huddle_rooms, 80),
        ("Phone Booths", program.support_spaces.phone_booths, 35),
        ("Break Rooms", program.support_spaces.break_rooms, 200),
        ("Reception Areas", program.support_spaces.reception_areas, 250),
        ("Copy/Print Centers", program.support_spaces.copy_print_centers, 100),
        ("Storage Rooms", program.support_spaces.storage_rooms, 150),
        ("Server/IT Rooms", program.support_spaces.server_rooms, 120),
        ("Wellness/Mother's Rooms", program.support_spaces.wellness_rooms, 80),
        ("Training Rooms", program.support_spaces.training_rooms, 600),
        ("Collaboration Areas", program.support_spaces.collaboration_areas, 200),
    ]
    
    row = 2
    for name, qty, sf_each in support_items:
        ws3[f'A{row}'] = name
        ws3[f'B{row}'] = qty
        ws3[f'B{row}'].font = input_font
        ws3[f'C{row}'] = sf_each
        ws3[f'D{row}'] = f'=B{row}*C{row}'
        ws3[f'D{row}'].number_format = '#,##0'
        row += 1
    
    # Total
    ws3[f'A{row}'] = "TOTAL SUPPORT SPACE"
    ws3[f'A{row}'].font = Font(bold=True)
    ws3[f'D{row}'] = f'=SUM(D2:D{row-1})'
    ws3[f'D{row}'].font = Font(bold=True)
    ws3[f'D{row}'].fill = total_fill
    ws3[f'D{row}'].number_format = '#,##0'
    
    for col, width in [(1, 35), (2, 12), (3, 12), (4, 15)]:
        ws3.column_dimensions[get_column_letter(col)].width = width
    
    # ========== REMOTE WORK ANALYSIS SHEET ==========
    ws4 = wb.create_sheet("Remote Work Analysis")
    
    ws4.merge_cells('A1:F1')
    ws4['A1'] = "REMOTE WORK IMPACT ANALYSIS"
    ws4['A1'].font = Font(bold=True, size=14)
    
    ws4['A3'] = f"Base Usable SF (Full On-Site):"
    ws4['B3'] = remote_analysis["base_usable_sf"]
    ws4['B3'].number_format = '#,##0'
    ws4['A4'] = f"Base Rentable SF (Full On-Site):"
    ws4['B4'] = remote_analysis["base_rentable_sf"]
    ws4['B4'].number_format = '#,##0'
    
    row = 6
    scenario_headers = ["Policy", "Description", "Reduction %", "Adj. Usable SF", "Adj. Rentable SF", "SF Saved"]
    for col, header in enumerate(scenario_headers, 1):
        cell = ws4.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    row = 7
    for scenario in remote_analysis["scenarios"]:
        ws4[f'A{row}'] = scenario["policy"].replace("_", " ").title()
        ws4[f'B{row}'] = scenario["description"]
        ws4[f'C{row}'] = scenario["percent_reduction"] / 100
        ws4[f'C{row}'].number_format = '0%'
        ws4[f'D{row}'] = scenario["adjusted_usable_sf"]
        ws4[f'D{row}'].number_format = '#,##0'
        ws4[f'E{row}'] = scenario["adjusted_rentable_sf"]
        ws4[f'E{row}'].number_format = '#,##0'
        ws4[f'F{row}'] = scenario["rentable_sf_saved"]
        ws4[f'F{row}'].number_format = '#,##0'
        row += 1
    
    # Recommendations
    row += 2
    ws4[f'A{row}'] = "RECOMMENDATIONS"
    ws4[f'A{row}'].font = header_font
    ws4[f'A{row}'].fill = header_fill
    ws4.merge_cells(f'A{row}:F{row}')
    row += 1
    
    for rec in remote_analysis["recommendations"]:
        ws4[f'A{row}'] = rec["category"]
        ws4[f'A{row}'].font = Font(bold=True)
        row += 1
        ws4[f'A{row}'] = rec["text"]
        ws4.merge_cells(f'A{row}:F{row}')
        row += 2
    
    for col, width in [(1, 20), (2, 35), (3, 12), (4, 15), (5, 15), (6, 12)]:
        ws4.column_dimensions[get_column_letter(col)].width = width
    
    # ========== SPACE STANDARDS REFERENCE SHEET ==========
    ws5 = wb.create_sheet("Space Standards")
    
    ws5['A1'] = "SPACE STANDARDS REFERENCE"
    ws5['A1'].font = Font(bold=True, size=14)
    
    std_headers = ["Space Type", "Standard SF", "Description"]
    for col, header in enumerate(std_headers, 1):
        cell = ws5.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    row = 4
    for key, std in DEFAULT_SPACE_STANDARDS.items():
        ws5[f'A{row}'] = std["name"]
        ws5[f'B{row}'] = std["sf"]
        ws5[f'C{row}'] = std["description"]
        row += 1
    
    for col, width in [(1, 25), (2, 12), (3, 40)]:
        ws5.column_dimensions[get_column_letter(col)].width = width
    
    wb.save(output_path)
    return output_path


def export_to_pdf(results: dict, remote_analysis: dict, program: SpaceProgram, output_path: str):
    """Export results to PDF report"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.units import inch
    
    doc = SimpleDocTemplate(output_path, pagesize=letter, 
                           leftMargin=0.75*inch, rightMargin=0.75*inch,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], 
                                  fontSize=18, spaceAfter=20, textColor=colors.HexColor('#2E5090'))
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], 
                                   fontSize=14, spaceBefore=15, spaceAfter=10, 
                                   textColor=colors.HexColor('#2E5090'))
    subheading_style = ParagraphStyle('SubHeading', parent=styles['Heading3'],
                                      fontSize=12, spaceBefore=10, spaceAfter=5)
    normal_style = styles['Normal']
    
    story = []
    
    # Title
    story.append(Paragraph("SPACE PROGRAMMING REPORT", title_style))
    story.append(Spacer(1, 10))
    
    # Company Info Table
    company_data = [
        ["Company:", results["company"]["name"]],
        ["Location:", results["company"]["location"]],
        ["Project:", results["company"]["project_name"]],
        ["Prepared By:", results["company"]["prepared_by"]],
        ["Date:", results["company"]["date"]],
    ]
    company_table = Table(company_data, colWidths=[1.5*inch, 4*inch])
    company_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(company_table)
    story.append(Spacer(1, 20))
    
    # Executive Summary
    story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
    
    summary_data = [
        ["Metric", "Value", "Unit"],
        ["Total Headcount", f"{results['totals']['total_staff']:,}", "persons"],
        ["Department Space", f"{results['totals']['department_sf']:,.0f}", "SF"],
        ["Support Space", f"{results['totals']['support_sf']:,.0f}", "SF"],
        ["Net Assignable SF", f"{results['totals']['net_assignable_sf']:,.0f}", "SF"],
        [f"Circulation ({results['totals']['circulation_factor']*100:.0f}%)", 
         f"{results['totals']['circulation_sf']:,.0f}", "SF"],
        ["Usable Square Feet", f"{results['totals']['usable_sf']:,.0f}", "SF"],
        ["", "", ""],
        ["Remote Work Policy", results['totals']['remote_work_description'], ""],
        ["Adjusted Usable SF", f"{results['totals']['adjusted_usable_sf']:,.0f}", "SF"],
        [f"Loss Factor ({results['totals']['loss_factor']*100:.0f}%)", 
         f"{results['totals']['loss_sf']:,.0f}", "SF"],
        ["RENTABLE SQUARE FEET", f"{results['totals']['rentable_sf']:,.0f}", "SF"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch, 1*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5090')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFC000')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Key Metrics
    story.append(Paragraph("KEY METRICS", heading_style))
    metrics_data = [
        ["Metric", "SF/Person"],
        ["Net Assignable", f"{results['metrics']['sf_per_person_net']:.1f}"],
        ["Usable", f"{results['metrics']['sf_per_person_usable']:.1f}"],
        ["Adjusted (w/ Remote)", f"{results['metrics']['sf_per_person_adjusted']:.1f}"],
        ["Rentable", f"{results['metrics']['sf_per_person_rentable']:.1f}"],
    ]
    metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(metrics_table)
    
    story.append(PageBreak())
    
    # Department Detail
    story.append(Paragraph("DEPARTMENT BREAKDOWN", heading_style))
    
    dept_data = [["Department", "Staff", "Total SF"]]
    for dept in results["departments"]:
        dept_data.append([dept["name"], str(dept["staff"]), f"{dept['breakdown']['total']:,.0f}"])
    dept_data.append(["TOTAL", str(results["totals"]["total_staff"]), 
                     f"{results['totals']['department_sf']:,.0f}"])
    
    dept_table = Table(dept_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    dept_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5090')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFC000')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(dept_table)
    story.append(Spacer(1, 20))
    
    # Remote Work Analysis
    story.append(Paragraph("REMOTE WORK IMPACT ANALYSIS", heading_style))
    story.append(Paragraph(
        f"Base Usable SF (Full On-Site): {remote_analysis['base_usable_sf']:,.0f} SF<br/>"
        f"Base Rentable SF (Full On-Site): {remote_analysis['base_rentable_sf']:,.0f} SF",
        normal_style))
    story.append(Spacer(1, 10))
    
    remote_data = [["Policy", "Description", "Reduction", "Adj. Rentable SF", "SF Saved"]]
    for scenario in remote_analysis["scenarios"]:
        remote_data.append([
            scenario["policy"].replace("_", " ").title(),
            scenario["description"][:30] + "..." if len(scenario["description"]) > 30 else scenario["description"],
            f"{scenario['percent_reduction']:.0f}%",
            f"{scenario['adjusted_rentable_sf']:,.0f}",
            f"{scenario['rentable_sf_saved']:,.0f}",
        ])
    
    remote_table = Table(remote_data, colWidths=[1.2*inch, 2*inch, 0.8*inch, 1.3*inch, 1*inch])
    remote_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(remote_table)
    story.append(Spacer(1, 15))
    
    # Recommendations
    story.append(Paragraph("RECOMMENDATIONS", subheading_style))
    for rec in remote_analysis["recommendations"]:
        story.append(Paragraph(f"<b>{rec['category']}:</b> {rec['text']}", normal_style))
        story.append(Spacer(1, 8))
    
    # Notes
    if results.get("notes"):
        story.append(PageBreak())
        story.append(Paragraph("NOTES", heading_style))
        story.append(Paragraph(results["notes"], normal_style))
    
    doc.build(story)
    return output_path


# ============================================================================
# INTERACTIVE CLI
# ============================================================================

def run_interactive():
    """Run interactive command-line interface"""
    print("\n" + "="*60)
    print("ARCHITECTURE SPACE PROGRAMMING TOOL")
    print("="*60)
    
    data_manager = DataManager()
    
    # Check for existing programs
    existing = data_manager.list_programs()
    if existing:
        print("\nExisting saved programs:")
        for i, prog in enumerate(existing, 1):
            print(f"  {i}. {prog['company_name']} - {prog['location']} ({prog['date_created']})")
        
        choice = input("\nLoad existing (enter number) or create new (press Enter)? ").strip()
        if choice.isdigit() and 0 < int(choice) <= len(existing):
            program = data_manager.load_program(existing[int(choice)-1]['filename'])
            print(f"\nLoaded: {program.company_name}")
        else:
            program = create_new_program()
    else:
        program = create_new_program()
    
    # Main menu
    while True:
        print("\n" + "-"*40)
        print("MAIN MENU")
        print("-"*40)
        print("1. View/Edit Company Information")
        print("2. Manage Departments")
        print("3. Manage Support Spaces")
        print("4. Set Factors (Circulation/Loss)")
        print("5. Set Remote Work Policy")
        print("6. Calculate & View Results")
        print("7. Export to PDF")
        print("8. Export to Excel")
        print("9. Save Program Data")
        print("10. Remote Work Analysis")
        print("0. Exit")
        
        choice = input("\nSelect option: ").strip()
        
        if choice == "1":
            edit_company_info(program)
        elif choice == "2":
            manage_departments(program)
        elif choice == "3":
            manage_support_spaces(program)
        elif choice == "4":
            set_factors(program)
        elif choice == "5":
            set_remote_policy(program)
        elif choice == "6":
            view_results(program)
        elif choice == "7":
            export_pdf(program, data_manager)
        elif choice == "8":
            export_excel(program, data_manager)
        elif choice == "9":
            filepath = data_manager.save_program(program)
            print(f"\nSaved to: {filepath}")
        elif choice == "10":
            view_remote_analysis(program)
        elif choice == "0":
            save = input("Save before exiting? (y/n): ").strip().lower()
            if save == 'y':
                data_manager.save_program(program)
            break


def create_new_program() -> SpaceProgram:
    """Create a new space program"""
    print("\n--- CREATE NEW PROGRAM ---")
    company = input("Company Name: ").strip()
    location = input("Location (City, State): ").strip()
    project = input("Project Name (optional): ").strip()
    prepared = input("Prepared By: ").strip()
    
    return SpaceProgram(
        company_name=company,
        location=location,
        project_name=project,
        prepared_by=prepared,
    )


def edit_company_info(program: SpaceProgram):
    """Edit company information"""
    print(f"\nCurrent Company: {program.company_name}")
    print(f"Current Location: {program.location}")
    print(f"Current Project: {program.project_name}")
    print(f"Prepared By: {program.prepared_by}")
    
    new_name = input(f"\nNew Company Name (Enter to keep): ").strip()
    if new_name:
        program.company_name = new_name
    
    new_loc = input(f"New Location (Enter to keep): ").strip()
    if new_loc:
        program.location = new_loc
    
    new_proj = input(f"New Project Name (Enter to keep): ").strip()
    if new_proj:
        program.project_name = new_proj
    
    new_prep = input(f"Prepared By (Enter to keep): ").strip()
    if new_prep:
        program.prepared_by = new_prep


def manage_departments(program: SpaceProgram):
    """Manage department staffing"""
    while True:
        print("\n--- DEPARTMENTS ---")
        if program.departments:
            for i, dept in enumerate(program.departments, 1):
                print(f"  {i}. {dept.name}: {dept.total_staff} staff")
        else:
            print("  No departments defined")
        
        print("\nOptions: (a)dd, (e)dit, (d)elete, (b)ack")
        choice = input("Choice: ").strip().lower()
        
        if choice == 'a':
            add_department(program)
        elif choice == 'e' and program.departments:
            idx = int(input("Department number to edit: ").strip()) - 1
            if 0 <= idx < len(program.departments):
                edit_department(program.departments[idx])
        elif choice == 'd' and program.departments:
            idx = int(input("Department number to delete: ").strip()) - 1
            if 0 <= idx < len(program.departments):
                del program.departments[idx]
        elif choice == 'b':
            break


def add_department(program: SpaceProgram):
    """Add a new department"""
    print("\n--- ADD DEPARTMENT ---")
    name = input("Department Name: ").strip()
    
    print("\nEnter staff counts by workspace type:")
    dept = Department(
        name=name,
        open_workstations=int(input("  Open Workstations (48 SF): ").strip() or 0),
        standard_workstations=int(input("  Standard Workstations (64 SF): ").strip() or 0),
        large_workstations=int(input("  Large Workstations (80 SF): ").strip() or 0),
        small_offices=int(input("  Small Private Offices (100 SF): ").strip() or 0),
        standard_offices=int(input("  Standard Private Offices (150 SF): ").strip() or 0),
        large_offices=int(input("  Large Private Offices (200 SF): ").strip() or 0),
        executive_offices=int(input("  Executive Offices (300 SF): ").strip() or 0),
    )
    program.departments.append(dept)
    print(f"\nAdded {name}: {dept.total_staff} total staff")


def edit_department(dept: Department):
    """Edit an existing department"""
    print(f"\n--- EDITING: {dept.name} ---")
    print("Enter new values (press Enter to keep current)")
    
    val = input(f"  Open Workstations [{dept.open_workstations}]: ").strip()
    if val:
        dept.open_workstations = int(val)
    
    val = input(f"  Standard Workstations [{dept.standard_workstations}]: ").strip()
    if val:
        dept.standard_workstations = int(val)
    
    val = input(f"  Large Workstations [{dept.large_workstations}]: ").strip()
    if val:
        dept.large_workstations = int(val)
    
    val = input(f"  Small Offices [{dept.small_offices}]: ").strip()
    if val:
        dept.small_offices = int(val)
    
    val = input(f"  Standard Offices [{dept.standard_offices}]: ").strip()
    if val:
        dept.standard_offices = int(val)
    
    val = input(f"  Large Offices [{dept.large_offices}]: ").strip()
    if val:
        dept.large_offices = int(val)
    
    val = input(f"  Executive Offices [{dept.executive_offices}]: ").strip()
    if val:
        dept.executive_offices = int(val)


def manage_support_spaces(program: SpaceProgram):
    """Manage support/amenity spaces"""
    ss = program.support_spaces
    print("\n--- SUPPORT SPACES ---")
    print("Enter quantities (press Enter to keep current)")
    
    val = input(f"  Small Conference Rooms (150 SF) [{ss.small_conference}]: ").strip()
    if val:
        ss.small_conference = int(val)
    
    val = input(f"  Medium Conference Rooms (300 SF) [{ss.medium_conference}]: ").strip()
    if val:
        ss.medium_conference = int(val)
    
    val = input(f"  Large Conference Rooms (500 SF) [{ss.large_conference}]: ").strip()
    if val:
        ss.large_conference = int(val)
    
    val = input(f"  Huddle Rooms (80 SF) [{ss.huddle_rooms}]: ").strip()
    if val:
        ss.huddle_rooms = int(val)
    
    val = input(f"  Phone Booths (35 SF) [{ss.phone_booths}]: ").strip()
    if val:
        ss.phone_booths = int(val)
    
    val = input(f"  Break Rooms (200 SF) [{ss.break_rooms}]: ").strip()
    if val:
        ss.break_rooms = int(val)
    
    val = input(f"  Reception Areas (250 SF) [{ss.reception_areas}]: ").strip()
    if val:
        ss.reception_areas = int(val)
    
    val = input(f"  Copy/Print Centers (100 SF) [{ss.copy_print_centers}]: ").strip()
    if val:
        ss.copy_print_centers = int(val)
    
    val = input(f"  Storage Rooms (150 SF) [{ss.storage_rooms}]: ").strip()
    if val:
        ss.storage_rooms = int(val)
    
    val = input(f"  Server/IT Rooms (120 SF) [{ss.server_rooms}]: ").strip()
    if val:
        ss.server_rooms = int(val)
    
    val = input(f"  Wellness/Mother's Rooms (80 SF) [{ss.wellness_rooms}]: ").strip()
    if val:
        ss.wellness_rooms = int(val)
    
    val = input(f"  Training Rooms (600 SF) [{ss.training_rooms}]: ").strip()
    if val:
        ss.training_rooms = int(val)
    
    val = input(f"  Collaboration Areas (200 SF) [{ss.collaboration_areas}]: ").strip()
    if val:
        ss.collaboration_areas = int(val)


def set_factors(program: SpaceProgram):
    """Set circulation and loss factors"""
    print("\n--- FACTORS ---")
    print(f"Current Circulation Factor: {program.circulation_factor*100:.0f}%")
    print(f"Current Loss Factor: {program.loss_factor*100:.0f}%")
    
    val = input(f"\nCirculation Factor % (industry standard 30-40%) [{program.circulation_factor*100:.0f}]: ").strip()
    if val:
        program.circulation_factor = float(val) / 100
    
    val = input(f"Loss Factor % for USF to RSF (typical 10-20%) [{program.loss_factor*100:.0f}]: ").strip()
    if val:
        program.loss_factor = float(val) / 100


def set_remote_policy(program: SpaceProgram):
    """Set remote work policy"""
    print("\n--- REMOTE WORK POLICY ---")
    print("Select policy:")
    for i, (key, info) in enumerate(REMOTE_WORK_FACTORS.items(), 1):
        marker = " <--" if key == program.remote_work_policy else ""
        print(f"  {i}. {key.replace('_', ' ').title()}: {info['description']}{marker}")
    
    choice = input("\nSelect (1-6): ").strip()
    if choice.isdigit() and 1 <= int(choice) <= 6:
        program.remote_work_policy = list(REMOTE_WORK_FACTORS.keys())[int(choice)-1]


def view_results(program: SpaceProgram):
    """Calculate and display results"""
    calc = SpaceCalculator(program)
    results = calc.calculate_totals()
    
    print("\n" + "="*60)
    print("SPACE PROGRAMMING RESULTS")
    print("="*60)
    print(f"Company: {results['company']['name']}")
    print(f"Location: {results['company']['location']}")
    print(f"Date: {results['company']['date']}")
    print("-"*60)
    
    print(f"\n{'TOTALS':^60}")
    print("-"*60)
    print(f"  Total Headcount:         {results['totals']['total_staff']:>15,} persons")
    print(f"  Department Space:        {results['totals']['department_sf']:>15,.0f} SF")
    print(f"  Support Space:           {results['totals']['support_sf']:>15,.0f} SF")
    print(f"  Net Assignable SF:       {results['totals']['net_assignable_sf']:>15,.0f} SF")
    print(f"  Circulation ({results['totals']['circulation_factor']*100:.0f}%):        {results['totals']['circulation_sf']:>15,.0f} SF")
    print(f"  Usable Square Feet:      {results['totals']['usable_sf']:>15,.0f} SF")
    print(f"\n  Remote Policy: {results['totals']['remote_work_description']}")
    print(f"  Adjusted Usable SF:      {results['totals']['adjusted_usable_sf']:>15,.0f} SF")
    print(f"  Loss Factor ({results['totals']['loss_factor']*100:.0f}%):         {results['totals']['loss_sf']:>15,.0f} SF")
    print("-"*60)
    print(f"  RENTABLE SQUARE FEET:    {results['totals']['rentable_sf']:>15,.0f} SF")
    print("="*60)
    
    print(f"\n{'KEY METRICS':^60}")
    print("-"*60)
    print(f"  SF per Person (Net):      {results['metrics']['sf_per_person_net']:>14.1f} SF")
    print(f"  SF per Person (Usable):   {results['metrics']['sf_per_person_usable']:>14.1f} SF")
    print(f"  SF per Person (Adjusted): {results['metrics']['sf_per_person_adjusted']:>14.1f} SF")
    print(f"  SF per Person (Rentable): {results['metrics']['sf_per_person_rentable']:>14.1f} SF")


def view_remote_analysis(program: SpaceProgram):
    """View remote work impact analysis"""
    analyzer = RemoteWorkAnalyzer(program)
    analysis = analyzer.analyze_scenarios()
    
    print("\n" + "="*60)
    print("REMOTE WORK IMPACT ANALYSIS")
    print("="*60)
    print(f"Base Usable SF (Full On-Site):   {analysis['base_usable_sf']:>12,.0f} SF")
    print(f"Base Rentable SF (Full On-Site): {analysis['base_rentable_sf']:>12,.0f} SF")
    print("-"*60)
    
    print(f"\n{'Policy':<20} {'Reduction':>10} {'Adj. RSF':>12} {'SF Saved':>12}")
    print("-"*60)
    for s in analysis['scenarios']:
        print(f"{s['policy'].replace('_',' ').title():<20} {s['percent_reduction']:>9.0f}% "
              f"{s['adjusted_rentable_sf']:>11,.0f} {s['rentable_sf_saved']:>11,.0f}")
    
    print("\n" + "-"*60)
    print("RECOMMENDATIONS:")
    for rec in analysis['recommendations']:
        print(f"\n{rec['category']}:")
        print(f"  {rec['text']}")


def export_pdf(program: SpaceProgram, data_manager: DataManager):
    """Export to PDF"""
    calc = SpaceCalculator(program)
    results = calc.calculate_totals()
    analyzer = RemoteWorkAnalyzer(program)
    analysis = analyzer.analyze_scenarios()
    
    safe_name = "".join(c if c.isalnum() else "_" for c in program.company_name)
    filename = f"{safe_name}_Space_Program.pdf"
    filepath = str(data_manager.data_dir / filename)
    
    export_to_pdf(results, analysis, program, filepath)
    print(f"\nPDF exported to: {filepath}")


def export_excel(program: SpaceProgram, data_manager: DataManager):
    """Export to Excel"""
    calc = SpaceCalculator(program)
    results = calc.calculate_totals()
    analyzer = RemoteWorkAnalyzer(program)
    analysis = analyzer.analyze_scenarios()
    
    safe_name = "".join(c if c.isalnum() else "_" for c in program.company_name)
    filename = f"{safe_name}_Space_Program.xlsx"
    filepath = str(data_manager.data_dir / filename)
    
    export_to_excel(results, analysis, program, filepath)
    print(f"\nExcel exported to: {filepath}")


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    run_interactive()
